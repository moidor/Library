import openpyxl
from src.streams.Database_text_streams.Authors_stream import authors_stream
from src.streams.Database_text_streams.Books_stream import books_stream
from src.streams.Database_text_streams.Customers_stream import customers_stream
from src.streams.Database_text_streams.Orders_stream import orders_stream
from src.streams.Database_text_streams.Sellers_stream import sellers_stream


workbook = openpyxl.load_workbook("C:\\Users\\Cham\\PycharmProjects\\Library\\src\\streams\\Tableur.xlsx")
sheets = workbook.sheetnames
# print(sheets)
# print(workbook.active.title)
authorsSheet = workbook["Authors"]
booksSheet = workbook["Books"]
customersSheet = workbook["Customers"]
ordersSheet = workbook["Orders"]
sellersSheet = workbook["Sellers"]


# value = authorsSheet["C4"].value
# print(authorsSheet)
# Tutorial: https://www.youtube.com/watch?v=XUYGxy4US7Q
# var = workbook["Authors"]["B2"].value
# print(var)
# print(authorsSheet.cell(2, 1).value)
# print(authorsSheet.cell(row=4, column=1).value)

class Database_Excel_stream:
    def iteration_excel_sheets(self):
        for sheet in sheets:
            # Itération sur toutes les feuilles
            if authorsSheet:
                rows = authorsSheet.max_row
                columns = authorsSheet.max_column
                # Itération sur toutes les lignes à partir de la deuxième pour ignorer les libellés de colonnes
                for row in range(2, rows + 1):  # rows: max_row + 1 pour itérer jusqu'à la dernière ligne
                    # Instanciation d'une nouvelle liste à chaque itération de ligne afin d'y insérer un (objet) auteur
                    author_list = list()
                    for column in range(1, columns + 1):
                        # Récupération du contenu de la cellule itérée
                        cell_content = authorsSheet.cell(row=row, column=column)
                        if cell_content.value is not None or "":  # Assurance que la cellule itérée soit bien renseignée
                            author_list.append(cell_content.value)  # Ajout de la cellule à la liste "author"
                            # IMPORTANT, on ne retient que les listes contenant les 6 arguments de l'objet "author"
                            if len(author_list) == 6:
                                print(author_list)
                                author_lastname, author_firstname, year_of_birth, city, country, language = \
                                    author_list[0], author_list[1], author_list[2], author_list[3], \
                                    author_list[4], author_list[5]
                                authors_stream(author_lastname, author_firstname, year_of_birth, city, country, language)

            if booksSheet:
                books_rows = booksSheet.max_row
                books_columns = booksSheet.max_column
                # Itération sur toutes les lignes à partir de la deuxième pour ignorer les libellés de colonnes
                for row in range(2, books_rows + 1):  # rows: max_row + 1 pour itérer jusqu'à la dernière ligne
                    # Instanciation d'une nouvelle liste à chaque itération de ligne afin d'y insérer un (objet) livre
                    book_list = list()
                    for column in range(1, books_columns + 1):
                        # Récupération du contenu de la cellule itérée
                        cell_content = booksSheet.cell(row=row, column=column)
                        book_list.append(cell_content.value)  # Ajout de la cellule à la liste "book"
                        # IMPORTANT, on ne retient que les listes contenant les 5 arguments de l'objet "book"
                        if len(book_list) == 5:
                            # print(book_list)
                            title, year, summary, author_firstname, author_lastname = \
                                book_list[0], book_list[1], book_list[2], book_list[3], book_list[4]
                            books_stream(title, year, summary, author_firstname, author_lastname)

            if customersSheet:
                customer_rows = customersSheet.max_row
                customer_columns = customersSheet.max_column
                # Itération sur toutes les lignes à partir de la deuxième pour ignorer les libellés de colonnes
                for row in range(2, customer_rows + 1):  # rows: max_row + 1 pour itérer jusqu'à la dernière ligne
                    # Instanciation d'une nouvelle liste à chaque itération de ligne afin d'y insérer un (objet) client
                    customer_list = list()
                    for column in range(1, customer_columns + 1):
                        # Récupération du contenu de la cellule itérée
                        cell_content = customersSheet.cell(row=row, column=column)
                        customer_list.append(cell_content.value)  # Ajout de la cellule à la liste "customer"
                        # IMPORTANT, on ne retient que les listes contenant les 8 arguments de l'objet "customer"
                        if len(customer_list) == 8:
                            print(customer_list)
                            lastname, firstname, year_of_birth, address, zip_code, city, country, language = \
                                customer_list[0], customer_list[1], customer_list[2], \
                                customer_list[3], customer_list[4], customer_list[5], customer_list[6], customer_list[7],
                            customers_stream(lastname, firstname, year_of_birth, address, zip_code, city, country, language)

            if ordersSheet:
                order_rows = ordersSheet.max_row
                order_columns = ordersSheet.max_column
                # Itération sur toutes les lignes à partir de la deuxième pour ignorer les libellés de colonnes
                for row in range(2, order_rows + 1):  # rows: max_row + 1 pour itérer jusqu'à la dernière ligne
                    # Instanciation d'une nouvelle liste à chaque itération de ligne afin d'y insérer un (objet)
                    # commande
                    order_list = list()
                    for column in range(1, order_columns + 1):
                        # Récupération du contenu de la cellule itérée
                        cell_content = ordersSheet.cell(row=row, column=column)
                        order_list.append(cell_content.value)  # Ajout de la cellule à la liste "order"
                        # IMPORTANT, on ne retient que les listes contenant les 5 arguments de l'objet "order"
                        if len(order_list) == 5:
                            print(order_list)
                            customer_lastname, customer_firstname, seller_name, book_title, date = \
                                order_list[0], order_list[1], order_list[2], order_list[3], order_list[4]
                            orders_stream(customer_lastname, customer_firstname, seller_name, book_title, date)

            if sellersSheet:
                seller_rows = sellersSheet.max_row
                seller_columns = sellersSheet.max_column
                # Itération sur toutes les lignes à partir de la deuxième pour ignorer les libellés de colonnes
                for row in range(2, seller_rows + 1):  # rows: max_row + 1 pour itérer jusqu'à la dernière ligne
                    # Instanciation d'une nouvelle liste à chaque itération de ligne afin d'y insérer un (objet) vendeur
                    seller_list = list()
                    for column in range(1, seller_columns + 1):
                        # Récupération du contenu de la cellule itérée
                        cell_content = sellersSheet.cell(row=row, column=column)
                        seller_list.append(cell_content.value)  # Ajout de la cellule à la liste "seller"
                        # IMPORTANT, on ne retient que les listes contenant les 3 arguments de l'objet "seller"
                        if len(seller_list) == 3:
                            print(seller_list)
                            seller_name, city, country = seller_list[0], seller_list[1], seller_list[2]
                            sellers_stream(seller_name, city, country)
