import openpyxl
from openpyxl.workbook import Workbook
from src.database.Connection import Connection
from src.interfaces.CountOfDatabase import CountOfDatabase


class DbBackup_Excel:
    def db_backup_excel(self):
        # Créer un fichier Excel
        workbook = Workbook()
        # Créer la feuille "Authors"
        workbook["Sheet"].title = "Authors"
        authors_sheet = workbook["Authors"]
        # active_sheet = workbook.active
        # Créer les titres des colonnes
        authors_sheet["A1"].value, authors_sheet["B1"].value, authors_sheet["C1"].value, \
            authors_sheet["D1"].value, authors_sheet["E1"].value, authors_sheet["F1"].value = \
            "Lastname", "firstname", "year_of_birth", "city", "country", "literature_language"

        command = f"""SELECT * FROM Authors;"""
        cursor = Connection.conn.execute(command)
        authors_db = cursor.fetchall()
        CountOfDatabase.count_of_database_method(authors_db)
        # Pour chaque auteur itéré, on ajoute une ligne dans le tableur avec les valeurs associées
        for author in authors_db:
            lastname, firstname, year_of_birth, city, country, literature_language \
                = author[1], author[2], author[3], author[4], author[5], author[6]
            # print(lastname, firstname, year_of_birth, city, country, literature_language)
            authors_sheet.append([lastname, firstname, year_of_birth, city, country, literature_language])

        # Créer la feuille "Books"
        workbook.create_sheet("Books")
        books_sheet = workbook["Books"]
        # Créer les titres des colonnes
        books_sheet["A1"].value, books_sheet["B1"].value, books_sheet["C1"].value = "Title", "Year", "Summary"

        command = f"""SELECT * FROM Books;"""
        cursor = Connection.conn.execute(command)
        books_db = cursor.fetchall()
        CountOfDatabase.count_of_database_method(books_db)
        # Pour chaque auteur itéré, on ajoute une ligne dans le tableur avec les valeurs associées
        for book in books_db:
            title, year, summary = \
                book[1], book[2], book[3]
            # print(title, year, summary, author_firstname, author_lastname)
            books_sheet.append([title, year, summary])

        # Créer la feuille "Customers"
        workbook.create_sheet("Customers")
        customers_sheet = workbook["Customers"]
        # Créer les titres des colonnes
        customers_sheet["A1"].value, customers_sheet["B1"].value, customers_sheet["C1"].value, \
            customers_sheet["D1"].value, customers_sheet["E1"].value, customers_sheet["F1"].value, \
            customers_sheet["G1"].value, customers_sheet["H1"].value = \
            "lastname", "firstname", "year_of_birth", "address", "zip_code", "city", "country", "language"

        command = f"""SELECT * FROM Customers;"""
        cursor = Connection.conn.execute(command)
        customers_db = cursor.fetchall()
        CountOfDatabase.count_of_database_method(customers_db)
        # Pour chaque client itéré, on ajoute une ligne dans le tableur avec les valeurs associées
        for customer in customers_db:
            lastname, firstname, year_of_birth, address, zip_code, city, country, language = \
                customer[1], customer[2], customer[3], customer[4], customer[5], customer[6], customer[7], customer[8]

            # print(title, year, summary, author_firstname, author_lastname)
            customers_sheet.append([lastname, firstname, year_of_birth, address, zip_code, city, country, language])

        # Créer la feuille "Orders"
        workbook.create_sheet("Orders")
        orders_sheet = workbook["Orders"]
        # Créer les titres des colonnes
        orders_sheet["A1"].value, orders_sheet["B1"].value, orders_sheet["C1"].value, orders_sheet["D1"].value = \
            "customerId", "sellerId", "bookId", "orderDate"

        command = f"""SELECT * FROM Orders;"""
        cursor = Connection.conn.execute(command)
        orders_db = cursor.fetchall()
        CountOfDatabase.count_of_database_method(orders_db)
        # Pour chaque client itéré, on ajoute une ligne dans le tableur avec les valeurs associées
        for order in orders_db:
            customerId, sellerId, bookId, orderDate = order[1], order[2], order[3], order[4]

            # print(title, year, summary, author_firstname, author_lastname)
            orders_sheet.append([customerId, sellerId, bookId, orderDate])

        # Créer la feuille "Sellers"
        workbook.create_sheet("Sellers")
        sellers_sheet = workbook["Sellers"]
        # Créer les titres des colonnes
        sellers_sheet["A1"].value, sellers_sheet["B1"].value, sellers_sheet["C1"].value = \
            "name", "city", "country"

        command = f"""SELECT * FROM Sellers;"""
        cursor = Connection.conn.execute(command)
        sellers_db = cursor.fetchall()
        CountOfDatabase.count_of_database_method(sellers_db)
        # Pour chaque client itéré, on ajoute une ligne dans le tableur avec les valeurs associées
        for seller in sellers_db:
            name, city, country = seller[1], seller[2], seller[3]

            # print(title, year, summary, author_firstname, author_lastname)
            sellers_sheet.append([name, city, country])

        # Sauvegarde du tableur
        workbook.save("C:\\Users\\Cham\\PycharmProjects\\Library\\src\\database\\excel_backup.xlsx")
